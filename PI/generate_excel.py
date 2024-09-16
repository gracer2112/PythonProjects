import pandas as pd
import json
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font

# Carregar dados das tarefas a partir do arquivo JSON
with open('data/tasks.json', encoding='utf-8') as f:
    tasks = json.load(f)

# Função para converter os dados em DataFrame
def tasks_to_dataframe(tasks, phase_name):
    data = []
    for task in tasks:
        data.append([
            task['id'],
            task['sequencia'],
            phase_name,
            task['atividade'],
            task['responsavel'],
            task['descricao'],
            task['data_inicio'],
            task['data_termino'],
            task.get('area', ''),  # Adiciona a área (Negócio ou TI)
            task.get('observacoes', '')  # Adiciona observações, se houver
        ])
    columns = ['#', 'Seq', 'Fase', 'Atividade', 'Responsável', 'Descrição', 'Data de Início', 'Data de Término', 'Área', 'Observações']
    return pd.DataFrame(data, columns=columns)

# Converter cada fase em DataFrame
df_pre_go_live = tasks_to_dataframe(tasks['pre_go_live'], 'Pré Go Live')
df_go_live = tasks_to_dataframe(tasks['go_live'], 'Go Live')
df_post_go_live = tasks_to_dataframe(tasks['post_go_live'], 'Pós Go Live')

# Adicionar linhas em branco entre as fases
blank_rows = pd.DataFrame([[''] * 10] * 3, columns=df_pre_go_live.columns)  # 3 linhas em branco

# Concatenar todos os DataFrames com linhas em branco entre as fases
df_all_tasks = pd.concat([df_pre_go_live, blank_rows, df_go_live, blank_rows, df_post_go_live], ignore_index=True)

# Exportar para Excel
excel_path = 'plano_de_implantacao.xlsx'
df_all_tasks.to_excel(excel_path, index=False, engine='openpyxl')

# Carregar o arquivo Excel e aplicar estilos
wb = load_workbook(excel_path)
ws = wb.active

# Definir estilos
negocio_fill = PatternFill(start_color="FFFF99", end_color="FFFF99", fill_type="solid")
ti_fill = PatternFill(start_color="CCFFCC", end_color="CCFFCC", fill_type="solid")
header_font = Font(bold=True)

# Aplicar estilos ao cabeçalho
for cell in ws[1]:
    cell.font = header_font

# Aplicar estilos às linhas com base na área
for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=10):
    area = row[8].value  # Coluna "Área"
    if area == "Negócio":
        for cell in row:
            cell.fill = negocio_fill
    elif area == "TI":
        for cell in row:
            cell.fill = ti_fill

# Salvar o arquivo Excel com os estilos aplicados
wb.save(excel_path)

print(f"Arquivo Excel '{excel_path}' gerado com sucesso e formatado com estilos!")