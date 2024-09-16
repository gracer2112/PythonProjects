from flask import  send_file, Blueprint, render_template, request, redirect, url_for, Flask,flash, abort
from datetime import datetime
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.worksheet.datavalidation import DataValidation
from collections import defaultdict
import os
from models import db, Project, Activity

project_bp = Blueprint('project_bp', __name__)

def map_values(activity):
    phase_mapping = {
        'pre_go_live': 'Pré Go Live',
        'go_live': 'Go Live',
        'post_go_live': 'Pós Go Live'
    }
    
    team_mapping = {
        'negocio': 'Negócio',
        'ti': 'TI'
    }
    
    activity.phase = phase_mapping.get(activity.phase, activity.phase)
    activity.equipe = team_mapping.get(activity.equipe, activity.equipe)
    return activity

@project_bp.route('/')
def list_projects():
    projects = Project.query.filter(Project.project_status != 'Concluido').all()
    return render_template('project_form.html', action='list', projects=projects)

@project_bp.route('/edit/<int:id>', methods=['GET', 'POST'])
def edit_project(id):
    project = Project.query.get_or_404(id)
    if request.method == 'POST':
        project.name = request.form['nome_projeto']
        project.meeting_date = datetime.strptime(request.form['meeting_date'], '%Y-%m-%d').date()
        project.document_version = request.form['document_version']
        
        db.session.commit()
        return redirect(url_for('project_bp.list_projects'))
    return render_template('project_form.html', action='edit', project=project)

@project_bp.route('/delete/<int:id>', methods=['POST'])
def delete_project(id):
    project = Project.query.get_or_404(id)
    
    # Verificar se há tarefas associadas ao projeto
    task_count = Activity.query.filter_by(projeto_id=id).count()
    
    if task_count > 0:
        flash('Não é possível excluir o projeto, pois há tarefas associadas a ele.', 'error')
    else:
        db.session.delete(project)
        db.session.commit()
        flash('Projeto excluído com sucesso.', 'success')
    
    return redirect(url_for('project_bp.list_projects'))

@project_bp.route('/generate_excel/<int:project_id>')
def generate_excel(project_id):
   
    # Consultar o nome do projeto usando o project_id
    project = Project.query.get(project_id)
    if not project:
        abort(404, description="Project not found.")
    project_name = project.nome_projeto


    # Consultar todas as atividades do projeto
    activities = Activity.query.filter_by(project_id=project_id).order_by(Activity.start_date, Activity.start_hour).all()
    activities = [map_values(activity) for activity in activities]

    # Organizar atividades por projeto, fase e equipe
    activities_by_project_phase_and_team = defaultdict(lambda: defaultdict(lambda: defaultdict(list)))
    for activity in activities:
        activities_by_project_phase_and_team[activity.project_id][activity.phase][activity.equipe].append(activity)

    # Criar um novo workbook e uma planilha
    wb = Workbook()
    ws = wb.active
    ws.title = "Plano de Implantação"

   # Definir estilos
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="333333", end_color="333333", fill_type="solid")
    project_fill = PatternFill(start_color="008000", end_color="008000", fill_type="solid")
    phase_fill = PatternFill(start_color="666666", end_color="666666", fill_type="solid")
    alignment = Alignment(horizontal="left", vertical="center")

    # Adicionar cabeçalho
    headers = ['#', 'Seq', 'Atividade', 'Responsável', 'Descrição', 'Data de Início', 'Hora de Início', 'Data de Término', 'Hora de Término', 'Status', 'Observações']
    ws.append(headers)

   # Aplicar estilo ao cabeçalho
    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = alignment


    # Adicionar dados ao Excel
    current_row = 2
    ws.cell(row=current_row, column=1).value = f"Projeto: {project_name} (ID: {project_id})"
    ws.cell(row=current_row, column=1).fill = project_fill
    ws.cell(row=current_row, column=1).font = Font(bold=True, color="FFFFFF")
    ws.cell(row=current_row, column=1).alignment = alignment
    ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=len(headers))
    current_row += 1

    for phase, teams in activities_by_project_phase_and_team[project_id].items():
        ws.cell(row=current_row, column=1).value = f"Fase: {phase}"
        ws.cell(row=current_row, column=1).fill = phase_fill
        ws.cell(row=current_row, column=1).font = Font(bold=True, color="FFFFFF")
        ws.cell(row=current_row, column=1).alignment = alignment
        ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=len(headers))
        current_row += 1

        for team, activities in teams.items():
            ws.cell(row=current_row, column=1).value = f"Equipe: {team}"
            ws.cell(row=current_row, column=1).fill = header_fill
            ws.cell(row=current_row, column=1).font = Font(bold=True, color="FFFFFF")
            ws.cell(row=current_row, column=1).alignment = alignment
            ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=len(headers))
            current_row += 1

            for activity in activities:
                ws.cell(row=current_row, column=1).value = activity.id           # #
                ws.cell(row=current_row, column=2).value = activity.seq           # Seq
                ws.cell(row=current_row, column=3).value =activity.activity      # Atividade
                ws.cell(row=current_row, column=4).value =activity.responsible   # Responsável
                ws.cell(row=current_row, column=5).value =activity.description  # Descrição
                ws.cell(row=current_row, column=6).value =activity.start_date    # Data de Início
                ws.cell(row=current_row, column=7).value = activity.start_hour    # Hora de Início
                ws.cell(row=current_row, column=8 ).value= activity.end_date      # Data de Término
                ws.cell(row=current_row, column=9).value= activity.end_hour      # Hora de Término
                ws.cell(row=current_row, column=10).value = ''                 # Status (pulada)
                ws.cell(row=current_row, column=11).value = activity.observations  # Observações
                current_row += 1
                            
    # Formatar colunas de data para dd-mm-yyyy
    date_columns = ['F', 'H']  # Colunas de 'Data de Início' e 'Data de Término'
    for col in date_columns:
        for cell in ws[col]:
            cell.number_format = 'DD-MM-YYYY'

    # Ajustar largura das colunas
    for col in ws.columns:
        max_length = max(len(str(cell.value)) for cell in col)
        adjusted_width = (max_length + 2)
        ws.column_dimensions[col[0].column_letter].width = adjusted_width

    # Salvar o arquivo Excel
    excel_path = os.path.join(os.getcwd(), 'plano_de_implantacao.xlsx')
    wb.save(excel_path)

    # Enviar o arquivo Excel como resposta
    return send_file(excel_path, as_attachment=True)
