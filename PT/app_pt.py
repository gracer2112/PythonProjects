import os
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from collections import defaultdict
import logging
import pandas as pd
from flask import Flask, render_template, request, redirect, url_for, flash, send_file, abort,jsonify
from models import db, TestCase, Funcionarios, Projeto  # Certifique-se de importar as classes necessárias
from datetime import datetime
from config import Config
from sqlalchemy.exc import SQLAlchemyError
from sqlalchemy import asc
from flask_migrate import Migrate

def create_app():
    app = Flask(__name__, template_folder='templates')
    app.config.from_object('config.Config')

    # Definindo uma chave secreta segura
    app.secret_key = os.urandom(24)    
    
    # Certifique-se de que o diretório 'instance' existe
    os.makedirs(os.path.join(Config.BASE_DIR, 'instance'), exist_ok=True)

    db.init_app(app)
    
    with app.app_context():
        db.create_all()

    migrate = Migrate(app, db)        
    
    # Rotas para CRUD
    @app.route('/')
    def index():
        project_id = request.args.get('project_id', type=int)
        clear_filters = request.args.get('clear_filters', False)      

        # Obter projetos que têm casos de teste
        projetos = Projeto.query.join(TestCase, Projeto.id == TestCase.id_projeto).distinct().all()

        # Se nenhum project_id for fornecido, redirecionar para o primeiro projeto com casos de teste
        if not project_id and not clear_filters and projetos:
            default_project_id = projetos[0].id  # Usar o primeiro projeto disponível
            return redirect(url_for('index', project_id=default_project_id))

        testcases = TestCase.query.filter_by(id_projeto=project_id).order_by(asc(TestCase.id_cenario), asc(TestCase.seq)).all() if project_id else []

        # Criar um dicionário para mapear IDs de funcionários para seus nomes e e-mails
        funcionarios = Funcionarios.query.all()
        #funcionarios_ti = Funcionario_TI.query.all()

        # Prefixar IDs para evitar conflitos
        funcionarios_dict = {funcionario.id: funcionario for funcionario in funcionarios}
        #funcionarios_dict.update({f'ti_{funcionario.ID}': funcionario for funcionario in funcionarios_ti})

        # Log para verificar IDs e funcionários carregados
        logging.info(f'Funcionários carregados: {funcionarios_dict.keys()}')
        for testcase in testcases:
            logging.info(f'Caso de Teste ID: {testcase.id}, Responsável ID: {testcase.id_responsavel}, Coordenador ID: {testcase.id_coordenador}')

        return render_template('index.html', testcases=testcases, projetos=projetos, funcionarios=funcionarios_dict, selected_project_id=project_id)

    @app.route('/create', methods=['GET', 'POST'])
    def create():
        if request.method == 'POST':
            try:
                new_case = TestCase(
                    seq=request.form['seq'],
                    dependencia=request.form['dependencia'],
                    id_cenario=request.form['id_cenario'],
                    modulo=request.form['modulo'],
                    caso_teste=request.form['caso_teste'],
                    info_teste=request.form['info_teste'],
                    passos=request.form['passos'],
                    resultado_esperado=request.form['resultado_esperado'],
                    status='Aguardando Início',
                    id_responsavel=request.form['responsavel'],
                    id_coordenador=request.form['coordenador'],
                    data_inicio_planejada=datetime.strptime(request.form['data_inicio_planejada'], '%Y-%m-%d'),
                    data_fim_planejada=datetime.strptime(request.form['data_fim_planejada'], '%Y-%m-%d'),
                    observacao=request.form['observacao'],
                    id_projeto=request.form['projeto']
                )
                db.session.add(new_case)
                db.session.commit()
                flash('Test case created successfully!', 'success')
                return redirect(url_for('index'))
            except (ValueError, SQLAlchemyError) as e:
                db.session.rollback()
                flash('Erro ao criar o caso de teste.', 'error')
                return render_template('create.html'), 400

        # Busca funcionários e projetos dentro do contexto da aplicação
        funcionarios = Funcionarios.query.order_by(Funcionarios.nome).all()
        projetos = Projeto.query.filter(Projeto.project_status != 'Concluido').order_by(Projeto.nome_projeto).all()

        # Calcula o próximo valor de seq
        max_seq = db.session.query(db.func.max(TestCase.seq)).scalar()
        next_seq = max_seq + 1 if max_seq is not None else 1

        return render_template('create.html', next_seq=next_seq, funcionarios=funcionarios, projetos=projetos)
    
    @app.route('/update/<int:id>', methods=['GET', 'POST'])
    def update(id):
        testcase = TestCase.query.get_or_404(id)

        if request.method == 'POST':
            responsavel_id = request.form['responsavel']
            
            try:
                testcase.seq = request.form['seq']
                testcase.dependencia = request.form['dependencia']
                testcase.id_cenario = request.form['id_cenario']
                testcase.modulo = request.form['modulo']
                testcase.caso_teste = request.form['caso_teste']
                testcase.info_teste = request.form['info_teste']
                testcase.passos = request.form['passos']
                testcase.resultado_esperado = request.form['resultado_esperado']
                new_status = request.form['status']

                # Atualizar status e datas conforme a regra de negócio
                if testcase.status == 'Aguardando Início' and new_status != 'Aguardando Início':
                    testcase.data_inicio_realizada = datetime.now()
                if new_status == 'Concluída':
                    testcase.data_fim_realizada = datetime.now()

                testcase.status = new_status
                testcase.id_responsavel = request.form['responsavel']
                testcase.id_coordenador = request.form['coordenador']
                # testcase.email_coordenador = request.form['email_coordenador']
                # testcase.email_responsavel = request.form['email_responsavel']
                testcase.data_inicio_planejada = datetime.strptime(request.form['data_inicio_planejada'], '%Y-%m-%d')
                testcase.data_fim_planejada = datetime.strptime(request.form['data_fim_planejada'], '%Y-%m-%d')
                testcase.data_inicio_realizada = datetime.strptime(request.form['data_inicio_realizada'], '%Y-%m-%d') if request.form['data_inicio_realizada'] else None
                testcase.data_fim_realizada = datetime.strptime(request.form['data_fim_realizada'], '%Y-%m-%d') if request.form['data_fim_realizada'] else None
                testcase.observacao = request.form['observacao']
                #testcase.id_projeto=request.form['projeto']                
                db.session.commit()
                flash('Test case updated successfully!', 'success')
                return redirect(url_for('index'))
            
            except (ValueError, SQLAlchemyError) as e:
                db.session.rollback()
                flash('Erro ao atualizar o caso de teste.', 'error')
                return render_template('update.html', testcase=testcase), 400

         # Busca funcionários e projetos
        funcionarios = Funcionarios.query.order_by(Funcionarios.nome).all()
        #funcionarios_ti = Funcionario_TI.query.order_by(Funcionario_TI.Nome).all()
        # funcionarios = Funcionarios.query.order_by(Funcionarios.nome).all()
        projetos = Projeto.query.all()

        # Prefixar IDs de funcionários de TI
        #funcionarios_ti_prefixed = [{'id': f'ti_{f.ID}', 'nome': f.Nome, 'email': f.Email} for f in funcionarios_ti]
        #funcionarios_prefixed = [{'id': f'gen_{f.id}', 'nome': f.nome, 'email': f.email} for f in funcionarios]

        #todos_funcionarios = funcionarios_prefixed + funcionarios_ti_prefixed

        return render_template('update.html', testcase=testcase, funcionarios=funcionarios, projetos=projetos)

    @app.route('/delete/<int:id>', methods=['GET', 'POST'])
    def delete(id):
        testcase = TestCase.query.get_or_404(id)
        if request.method == 'POST':
            try:
                db.session.delete(testcase)
                db.session.commit()
                flash('Test case deleted successfully!', 'success')
            except SQLAlchemyError as e:
                db.session.rollback()
                flash('Erro ao deletar o caso de teste.', 'error')
                return render_template('delete.html', testcase=testcase), 400
            return redirect(url_for('index'))
        return render_template('delete.html', testcase=testcase)

    @app.route('/generate_excel/<int:project_id>', methods=['POST'])
    def generate_excel(project_id):
        data = request.json.get('data', [])
        
        if not data:
            abort(400, description="No data provided.")

        # Criar um novo workbook e uma planilha
        wb = Workbook()
        ws = wb.active
        ws.title = "Plano de Teste"

        # Definir cabeçalhos
        headers = [
            "ID", "Seq", "Dependência", "ID do Cenário", "Módulo", "Caso de Teste",
            "Informações para o Teste", "Passos", "Resultado Esperado", "Status",
            "Responsável", "Email do Responsável", "Coordenador", "Email do Coordenador",
            "Data Início Planejada", "Data Fim Planejada", "Data Início Realizada",
            "Data Fim Realizada", "Observação"
        ]
        ws.append(headers)

        # Definir estilos
        header_fill = PatternFill(start_color="007a3d", end_color="007a3d", fill_type="solid")
        status_fill = PatternFill(start_color="FFA500", end_color="FFA500", fill_type="solid")  # Laranja
        gray_fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")  # Cinza
        thin_border = Border(left=Side(style='thin'), right=Side(style='thin'),
                            top=Side(style='thin'), bottom=Side(style='thin'))

        # Aplicar estilos aos cabeçalhos
        for col in ws.iter_cols(min_row=1, max_row=1, min_col=1, max_col=len(headers)):
            for cell in col:
                cell.font = Font(bold=True, color="FFFFFF")
                cell.alignment = Alignment(horizontal='center')
                cell.fill = header_fill
                cell.border = thin_border

        # Adicionar dados e aplicar estilos
        for row_data in data:
            ws.append(row_data)

        # Aplicar bordas e estilos às colunas específicas
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=len(headers)):
            for cell in row:
                # Aplicar estilo laranja à coluna "Status"
                if cell.column == 10:  # Coluna "Status"
                    cell.fill = status_fill
                # Aplicar estilo cinza às colunas "Informações para o Teste", "Passos" e "Resultado Esperado"
                elif cell.column in [7, 8, 9]:  # Colunas "Informações para o Teste", "Passos", "Resultado Esperado"
                    cell.fill = gray_fill
                cell.border = thin_border

        # Salvar o arquivo
        file_path = os.path.join(os.getcwd(),'test_cases.xlsx')
        wb.save(file_path)

        # Enviar arquivo para download
        return send_file(file_path, as_attachment=True)
    
    @app.route('/upload_file', methods=['POST'])
    def upload_file():
        file = request.files.get('file')
        if not file:
            flash('Nenhum arquivo selecionado.', 'error')
            return redirect(url_for('index'))

        try:
            # Ler o arquivo Excel
            df = pd.read_excel(file)

            # Iterar sobre as linhas do DataFrame
            for index, row in df.iterrows():
                testcase_id = row['ID']
                testcase = TestCase.query.get(testcase_id)

                if testcase:
                    # Atualizar os campos necessários
                    testcase.status = row['Status']
                    testcase.data_inicio_realizada = pd.to_datetime(row['Data Início Realizada']) if not pd.isna(row['Data Início Realizada']) else None
                    testcase.data_fim_realizada = pd.to_datetime(row['Data Fim Realizada']) if not pd.isna(row['Data Fim Realizada']) else None
                    # Adicione outras atualizações de campo conforme necessário

            db.session.commit()
            flash('Casos de teste atualizados com sucesso!', 'success')
        except Exception as e:
            db.session.rollback()
            flash('Erro ao processar o arquivo.', 'error')
            logging.error(f'Erro ao atualizar casos de teste: {e}')
        return redirect(url_for('index'))  

    @app.route('/upload_test_cases', methods=['POST'])
    def upload_test_cases():
        # Verifica se um projeto está selecionado
        project_id = request.form.get('projeto')
        if not project_id:
            return jsonify(success=False, message='Por favor, selecione um projeto.')

        file = request.files.get('file')
        if not file:
            return jsonify(success=False, message='Nenhum arquivo selecionado.')


        try:
            # Ler o arquivo Excel
            df = pd.read_excel(file)

            # Defina as colunas esperadas
            colunas_esperadas = [
                'Seq', 'Dependência', 'ID do Cenário', 'Módulo', 'Caso de Teste',
                'Informações para o Teste', 'Passos', 'Resultado Esperado',
                'Email do Responsável', 'Email do Coordenador', 'Data Início Planejada',
                'Data Fim Planejada', 'Observação'
            ]

            colunas_faltando = [col for col in colunas_esperadas if col not in df.columns]
            if colunas_faltando:
                return jsonify(success=False, message=f'Faltando colunas: {", ".join(colunas_faltando)}')
            
            # Criar um dicionário de e-mails para IDs de funcionários
            funcionarios_dict = {funcionario.email: funcionario.id for funcionario in Funcionarios.query.all()}

            # Iterar sobre as linhas do DataFrame
            for index, row in df.iterrows():
                # Verifica se todos os campos obrigatórios estão preenchidos
                if pd.isna(row['Seq']) or pd.isna(row['Dependência']) or pd.isna(row['ID do Cenário']) or pd.isna(row['Módulo']) or pd.isna(row['Caso de Teste']) or pd.isna(row['Informações para o Teste']) or pd.isna(row['Passos']) or pd.isna(row['Resultado Esperado']):
                    continue

                # Obter IDs de responsável e coordenador usando e-mails
                id_responsavel = funcionarios_dict.get(row['Email do Responsável'])
                id_coordenador = funcionarios_dict.get(row['Email do Coordenador'])

                # Verifica se os IDs foram encontrados
                if not id_responsavel or not id_coordenador:
                    flash(f'Erro: Não foi possível encontrar responsável ou coordenador para o e-mail na linha {index+1}.', 'error')
                    continue

                # Cria um novo caso de teste
                new_case = TestCase(
                    seq=row['Seq'],
                    dependencia=row['Dependência'],
                    id_cenario=row['ID do Cenário'],
                    modulo=row['Módulo'],
                    caso_teste=row['Caso de Teste'],
                    info_teste=row['Informações para o Teste'],
                    passos=row['Passos'],
                    resultado_esperado=row['Resultado Esperado'],
                    status='Aguardando Início',
                    id_responsavel=id_responsavel,
                    id_coordenador=id_coordenador,
                    data_inicio_planejada=pd.to_datetime(row['Data Início Planejada']) if not pd.isna(row['Data Início Planejada']) else None,
                    data_fim_planejada=pd.to_datetime(row['Data Fim Planejada']) if not pd.isna(row['Data Fim Planejada']) else None,
                    observacao=row['Observação'] if not pd.isna(row['Observação']) else '',
                    id_projeto=project_id
                )
                db.session.add(new_case)

            db.session.commit()
            return jsonify(success=True)
        except Exception as e:
            logging.error(f'Erro ao processar o arquivo: {e}')
            return jsonify(success=False, message='Erro ao processar o arquivo.')

    return app

if __name__ == '__main__':
    app = create_app()
    app.run(debug=True,port=5002)